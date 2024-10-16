"""
Import risk excel 🕸️ into RiskMatrix ✨

This script is specific to our sitation at seantis. The script is included
anyway, you might adjust it to import the excel at your organization too.
"""
import argparse
import sys
import traceback
from datetime import datetime
from typing import TYPE_CHECKING
from typing import Iterator

try:
    from openpyxl import load_workbook
except ImportError:
    print("Excel import requires openpyxl library. Install with:\n")
    print("$ pip install openpyxl")
    print()
    sys.exit(1)

from riskmatrix.models.risk_assessment_info import RiskAssessmentInfo, RiskAssessmentState
import sqlalchemy
from pyramid.paster import bootstrap
from pyramid.paster import get_appsettings
from sqlalchemy import select

from riskmatrix.models import Asset
from riskmatrix.models import Organization
from riskmatrix.models import Risk
from riskmatrix.models import RiskAssessment
from riskmatrix.models import RiskCatalog
from riskmatrix.orm import Base
from riskmatrix.orm import get_engine
from riskmatrix.scripts.util import select_existing_organization

if TYPE_CHECKING:

    from typing import TypedDict

    from sqlalchemy.orm import Session

    class RiskDetails(TypedDict):
        """ A risk extracted from the excel. """
        name: str
        category: str
        asset_name: str
        desc: str
        likelihood: int
        impact: int


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        'config_uri',
        help='Configuration file, e.g., development.ini',
    )
    parser.add_argument(
        'catalog',
        help='Risk catalog excel file, e.g., catalog.xlsx',
    )
    return parser.parse_args(argv[1:])


def get_or_create_asset(
    asset_name: str,
    organization: Organization,
    session: 'Session'
) -> Asset:

    q = select(Asset).where(
        Asset.organization_id == organization.id,
        Asset.name == asset_name
    )

    if asset := session.scalars(q).one_or_none():
        return asset

    asset = Asset(asset_name, organization)
    asset.organization_id = organization.id
    session.add(asset)
    return asset


def get_or_create_risk(
    risk_name: str,
    catalog: RiskCatalog,
    session: 'Session'
) -> Risk:

    q = select(Risk).where(
        Risk.organization_id == catalog.organization.id,
        Risk.name == risk_name
    )

    if risk := session.scalars(q).one_or_none():
        return risk

    risk = Risk(risk_name, catalog)
    session.add(risk)
    return risk


def get_or_create_risk_assessment(
    risk: Risk,
    asset: Asset,
    organization: Organization,
    session: 'Session'
) -> RiskAssessment:

    q = select(RiskAssessment).where(
        RiskAssessment.risk_id == risk.id,
        RiskAssessment.asset_id == asset.id,
    )

    if assessment := session.scalars(q).one_or_none():
        return assessment
    
    # check if theres an existing instance of RiskAssessmentInfo which is open for the organization
    q = select(RiskAssessmentInfo).where(
        RiskAssessmentInfo.organization_id == organization.id,
        RiskAssessmentInfo.state == RiskAssessmentState.OPEN
    )
    if risk_assessment_info := session.scalars(q).one_or_none():
        assessment = RiskAssessment(risk=risk, asset=asset, info=risk_assessment_info)
        session.add(assessment)
        return assessment
    
    risk_assessment_info = RiskAssessmentInfo(organization_id=organization.id)
    
    risk_assessment_info.state = RiskAssessmentState.OPEN
    risk_assessment_info.organization_id = organization.id
    
    session.add(risk_assessment_info)
    # flush
    session.flush()
    session.refresh(risk_assessment_info)
    

    assessment = RiskAssessment(risk=risk, asset=asset, info=risk_assessment_info)
    session.add(assessment)
    return assessment


def populate_catalog(
    catalog: RiskCatalog,
    risks: 'Iterator[RiskDetails]',
    session: 'Session'
) -> None:

    for risk_details in risks:
        asset = get_or_create_asset(
            risk_details['asset_name'], catalog.organization, session
        )

        risk = get_or_create_risk(
            risk_details['name'], catalog, session
        )
        risk.category = risk_details['category']
        risk.description = risk_details['desc']

        assessment = get_or_create_risk_assessment(risk, asset, catalog.organization, session)
        assessment.likelihood = risk_details['likelihood']
        assessment.impact = risk_details['impact']


def risks_from_excel(
    excel_file: str,
    sheet_name: str = 'Risikokatalog'
) -> 'Iterator[RiskDetails]':
    """
    Load risks from excel.
    """
    workbook = load_workbook(excel_file, read_only=True)

    sheet = workbook[sheet_name]

    # Rows are vertically grouped into sections by a category. A section begins
    # with a row that contains the category name but is otherwise empty.
    current_category = None

    # Header row sometimes spans over two rows (combined), sometimes only one.
    # Anyway, actual riks rows will start after row #2.
    start_after_row = 2

    iterator = sheet.iter_rows(
        values_only=True,
        min_row=start_after_row
    )

    for row in iterator:
        nr = row[0]
        name = row[1]

        is_empty_row = not (nr or name)
        is_category_row = not nr and name

        if is_empty_row:
            continue
        elif is_category_row:
            current_category = name
            continue

        yield {
            'name': str(name),
            'category': str(current_category),
            'asset_name': str(row[2]),
            'desc': str(row[3]),
            'likelihood': int(str(row[7])),
            'impact': int(str(row[8]))
        }

    # readonly mode forces us to manually close the workbook, see also:
    # https://openpyxl.readthedocs.io/en/stable/optimized.html#read-only-mode
    workbook.close()


def main(argv: list[str] = sys.argv) -> None:
    args = parse_args(argv)

    with bootstrap(args.config_uri) as env:
        settings = get_appsettings(args.config_uri)

        engine = get_engine(settings)
        Base.metadata.create_all(engine)

        with env['request'].tm:
            dbsession = env['request'].dbsession

            print('Organization to attach risk catalog to')

            org = select_existing_organization(dbsession)

            if not org:
                return

            today = datetime.today().strftime('%Y-%m-%d')

            catalog = RiskCatalog(
                'seantis risk register',
                organization=org,
                description=f'Imported from risk excel on {today}.'
            )

            catalog.organization_id = org.id

            try:
                populate_catalog(
                    catalog,
                    risks_from_excel(args.catalog),
                    dbsession
                )
            except sqlalchemy.exc.IntegrityError:
                print('Failed to import excel, aborting.')
                print(traceback.format_exc())
                dbsession.rollback()
                sys.exit(1)
            else:
                print(
                    f'Successfully populated risk catalog "{catalog.name}" '
                    'from risk register excel.'
                )


if __name__ == '__main__':
    main(sys.argv)
